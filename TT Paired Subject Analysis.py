# Analyzes two subfolders with the SAME SUBJECTS in both folders. Paired differences analysis.
# Generates an excel file with all the useful statistical information.

import os
import pandas as pd
import numpy as np
from scipy.stats import ttest_rel, wilcoxon, shapiro
from statsmodels.stats.power import TTestPower
from statsmodels.stats.multitest import multipletests
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill

# Create and hide the Tkinter root window.
root = tk.Tk()
root.withdraw()

def auto_adjust_and_center_and_highlight(filename):
    wb = load_workbook(filename)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for ws in wb.worksheets:
        # Adjust column widths and center text.
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width
            if col_letter != "A":
                for cell in col:
                    cell.alignment = Alignment(horizontal="center")

        # Find the column index for "Significance" header.
        significance_col = None
        for col in ws.iter_cols(min_row=1, max_row=1):
            header = str(col[0].value).strip().lower()
            if header == "significance":
                significance_col = col[0].column
                break
        if significance_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[significance_col - 1]
                try:
                    if float(cell.value) or cell.value:
                        # If the cell's text (uppercased) equals "SIGNIFICANT", highlight the row.
                        if str(cell.value).strip().upper() == "SIGNIFICANT":
                            for c in row:
                                c.fill = highlight_fill
                except Exception:
                    continue
    wb.save(filename)

def extract_subject_id(filename):
    base = os.path.splitext(filename)[0]
    tokens = base.split()
    return tokens[0] if tokens else base

def read_folder_data(folder_path):
    """
    Read all CSV and Excel (.xlsx) files in a folder that start with a capital "P".
    For each file, the header row (cells A1 and B1) is used to define the column names.
    Returns a tuple: (combined DataFrame, variable_order_list)
    """
    data_dict = {}
    variable_order = None
    for filename in os.listdir(folder_path):
        if not (filename.startswith("P") and (filename.lower().endswith('.csv') or filename.lower().endswith('.xlsx'))):
            continue
        full_path = os.path.join(folder_path, filename)
        try:
            if filename.lower().endswith('.csv'):
                df = pd.read_csv(full_path)
            elif filename.lower().endswith('.xlsx'):
                df = pd.read_excel(full_path)
        except Exception as e:
            print(f"Error reading {filename}: {e}")
            continue
        # Use the header values from cell A1 and B1 as the column names.
        var_col_name = df.columns[0]
        val_col_name = df.columns[1]
        # Standardize variable names.
        df[var_col_name] = df[var_col_name].astype(str).str.lower().str.strip()
        if variable_order is None:
            variable_order = list(df[var_col_name])
        series = df.set_index(var_col_name)[val_col_name].apply(pd.to_numeric, errors='coerce')
        subject_id = extract_subject_id(filename)
        data_dict[subject_id] = series
    if data_dict:
        return pd.DataFrame.from_dict(data_dict, orient='index'), variable_order
    else:
        return None, []

# --- Main Script ---

# Dynamically select two folders.
folder1 = filedialog.askdirectory(title="Select Folder 1")
if not folder1:
    print("No Folder 1 selected. Exiting...")
    exit()
folder2 = filedialog.askdirectory(title="Select Folder 2")
if not folder2:
    print("No Folder 2 selected. Exiting...")
    exit()

folder1_name = os.path.basename(os.path.normpath(folder1))
folder2_name = os.path.basename(os.path.normpath(folder2))

# Read data from both folders.
df1, var_order1 = read_folder_data(folder1)
df2, var_order2 = read_folder_data(folder2)

if df1 is None or df2 is None:
    print("One or both folders could not be processed. Exiting...")
    exit()

# Merge the two DataFrames by subject ID.
common_subjects = list(set(df1.index).intersection(df2.index))
print(f"Common subjects found: {len(common_subjects)}")
if not common_subjects:
    print("No common subjects found. Exiting...")
    exit()
df1 = df1.loc[common_subjects]
df2 = df2.loc[common_subjects]

# Determine common variables across both folders.
common_vars = list(set(df1.columns).intersection(df2.columns))
print(f"Common variables found: {len(common_vars)}")
if not common_vars:
    print("No common variables found. Exiting...")
    exit()

if var_order1:
    variable_order = [var for var in var_order1 if var in common_vars]
else:
    variable_order = common_vars

# Prepare list for storing comparison results.
results = []
power_estimator = TTestPower()

for var in variable_order:
    folder1_series = pd.to_numeric(df1[var], errors='coerce')
    folder2_series = pd.to_numeric(df2[var], errors='coerce')
    paired_data = pd.DataFrame({'Folder1': folder1_series, 'Folder2': folder2_series}).dropna()
    if paired_data.shape[0] < 3:
        continue
    folder1_avg = paired_data['Folder1'].mean()
    folder2_avg = paired_data['Folder2'].mean()
    differences = paired_data['Folder1'] - paired_data['Folder2']
    n = len(differences)

    try:
        norm_stat, norm_p = shapiro(differences)
    except Exception as e:
        print(f"Error performing Shapiro test on {var}: {e}")
        continue

    if norm_p > 0.05:
        test_used = "Paired t-test"
        try:
            t_stat, p_value = ttest_rel(paired_data['Folder1'], paired_data['Folder2'])
        except Exception as e:
            print(f"Error performing t-test on {var}: {e}")
            continue
        std_diff = differences.std(ddof=1)
        cohen_d = differences.mean() / std_diff if std_diff != 0 else np.nan
        try:
            power = power_estimator.power(effect_size=abs(cohen_d), nobs=n, alpha=0.05, alternative='two-sided')
        except Exception:
            power = np.nan
    else:
        test_used = "Wilcoxon signed-rank test"
        try:
            w_stat, p_value = wilcoxon(paired_data['Folder1'], paired_data['Folder2'])
        except Exception as e:
            print(f"Error performing Wilcoxon test on {var}: {e}")
            continue
        expected_W = n * (n + 1) / 4
        std_W = np.sqrt(n * (n + 1) * (2 * n + 1) / 24)
        z_val = (w_stat - expected_W) / std_W if std_W != 0 else np.nan
        cohen_d = abs(z_val) / np.sqrt(n) if n > 0 else np.nan
        power = np.nan

    significance = "SIGNIFICANT" if p_value < 0.05 else "NOT SIGNIFICANT"

    results.append({
        "Variable": var,
        f"{folder1_name} Avg": folder1_avg,
        f"{folder2_name} Avg": folder2_avg,
        "Test Used": test_used,
        "p-value": p_value,
        "Effect Size": cohen_d,
        "Power": power,
        "Significance": significance
    })

results_df = pd.DataFrame(results)

default_filename = f"{folder1_name} vs {folder2_name} Comparison.xlsx"
# Perform FDR correction
p_values = results_df["p-value"].values
_, corrected_pvals, _, _ = multipletests(p_values, method='fdr_bh')
results_df["FDR Corrected p-value"] = corrected_pvals
results_df["FDR Significance"] = ["SIGNIFICANT" if p < 0.05 else "NOT SIGNIFICANT" for p in corrected_pvals]

# Save the updated dataframe to Excel
output_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                           initialfile=default_filename,
                                           title="Save comparison results as",
                                           filetypes=[("Excel Files", "*.xlsx")])
if output_path:
    try:
        results_df.to_excel(output_path, index=False)
        auto_adjust_and_center_and_highlight(output_path)
        print(f"Comparison results saved to {output_path}")
    except Exception as e:
        print(f"Error saving comparison results: {e}")
else:
    print("No save location selected. Exiting...")
