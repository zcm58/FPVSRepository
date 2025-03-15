# Analyzes unbalanced TT Serve vs Return Datasets using Linear Mixed Effects Model
#

import os
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
import statsmodels.formula.api as smf

# Create and hide the Tkinter root window.
root = tk.Tk()
root.withdraw()


def auto_adjust_and_center_and_highlight(filename):
    wb = load_workbook(filename)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width
            if col_letter != "A":
                for cell in col:
                    cell.alignment = Alignment(horizontal="center")

        significance_col = None
        for col in ws.iter_cols(min_row=1, max_row=1):
            header = str(col[0].value).strip().lower()
            if header == "significance":
                significance_col = col[0].column
                break
        if significance_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[significance_col - 1]
                try:
                    if str(cell.value).strip().upper() == "SIGNIFICANT":
                        for c in row:
                            c.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                except Exception:
                    continue
    wb.save(filename)


def extract_subject_id(filename):
    base = os.path.splitext(filename)[0]
    tokens = base.split()
    return tokens[0] if tokens else base


def read_folder_data(folder_path):
    data_dict = {}
    variable_order = None
    for filename in os.listdir(folder_path):
        if not (filename.startswith("P") and filename.lower().endswith('.csv')):
            continue
        full_path = os.path.join(folder_path, filename)
        try:
            df = pd.read_csv(full_path)
            var_col_name = df.columns[0]
            val_col_name = df.columns[1]
            df[var_col_name] = df[var_col_name].astype(str).str.lower().str.strip()
            if variable_order is None:
                variable_order = list(df[var_col_name])
            series = df.set_index(var_col_name)[val_col_name].apply(pd.to_numeric, errors='coerce')
            subject_id = extract_subject_id(filename)
            data_dict[subject_id] = series
        except Exception as e:
            print(f"Error reading {filename}: {e}")
    if data_dict:
        return pd.DataFrame.from_dict(data_dict, orient='index'), variable_order
    else:
        return None, []


# --- Main Script ---

folder1 = filedialog.askdirectory(title="Select Folder 1")
if not folder1:
    print("No Folder 1 selected. Exiting...")
    exit()
folder2 = filedialog.askdirectory(title="Select Folder 2")
if not folder2:
    print("No Folder 2 selected. Exiting...")
    exit()

folder1_name = os.path.basename(os.path.normpath(folder1))
folder2_name = os.path.basename(os.path.normpath(folder2))

df1, var_order1 = read_folder_data(folder1)
df2, var_order2 = read_folder_data(folder2)

if df1 is None or df2 is None:
    print("One or both folders could not be processed. Exiting...")
    exit()

common_vars = list(set(df1.columns).intersection(df2.columns))
if not common_vars:
    print("No common variables found. Exiting...")
    exit()

variable_order = [var for var in var_order1 if var in common_vars] if var_order1 else common_vars

combined_df = pd.concat([
    df1[common_vars].assign(Task=folder1_name),
    df2[common_vars].assign(Task=folder2_name)
]).reset_index().rename(columns={'index': 'Subject'})

results = []

for var in variable_order:
    analysis_df = combined_df[['Subject', 'Task', var]].dropna().rename(columns={var: 'Value'})

    if analysis_df['Task'].nunique() < 2:
        continue

    model = smf.mixedlm("Value ~ Task", data=analysis_df, groups=analysis_df["Subject"])
    try:
        result = model.fit()
        p_value = result.pvalues.get('Task[T.' + folder2_name + ']', np.nan)
        coef = result.params.get('Task[T.' + folder2_name + ']', np.nan)
    except Exception as e:
        print(f"Error fitting LMM for {var}: {e}")
        continue

    significance = "SIGNIFICANT" if p_value < 0.05 else "NOT SIGNIFICANT"

    results.append({
        "Variable": var,
        f"{folder1_name} Avg": analysis_df[analysis_df['Task'] == folder1_name]['Value'].mean(),
        f"{folder2_name} Avg": analysis_df[analysis_df['Task'] == folder2_name]['Value'].mean(),
        "Coefficient": coef,
        "p-value": p_value,
        "Significance": significance
    })

results_df = pd.DataFrame(results)

default_filename = f"{folder1_name} vs {folder2_name} LMM Comparison.xlsx"
output_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                           initialfile=default_filename,
                                           title="Save comparison results as",
                                           filetypes=[("Excel Files", "*.xlsx")])
if output_path:
    try:
        results_df.to_excel(output_path, index=False)
        auto_adjust_and_center_and_highlight(output_path)
        print(f"Comparison results saved to {output_path}")
    except Exception as e:
        print(f"Error saving comparison results: {e}")
else:
    print("No save location selected. Exiting...")