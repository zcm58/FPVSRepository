# Averages the variables of two subfolders together to create an average of them both.

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import os
import openpyxl
import csv
import re
import sys
import subprocess
from openpyxl.styles import Alignment


def select_folder1():
    folder = filedialog.askdirectory()
    if folder:
        folder1_entry.delete(0, tk.END)
        folder1_entry.insert(0, folder)


def select_folder2():
    # Open Folder 2 dialog in the parent directory of Folder 1 (not inside Folder 1)
    folder1 = folder1_entry.get()
    if folder1:
        initial_dir = os.path.dirname(folder1)
    else:
        initial_dir = None
    folder = filedialog.askdirectory(initialdir=initial_dir)
    if folder:
        folder2_entry.delete(0, tk.END)
        folder2_entry.insert(0, folder)


def select_save_folder():
    folder = filedialog.askdirectory()
    if folder:
        save_folder_entry.delete(0, tk.END)
        save_folder_entry.insert(0, folder)


def read_workbook_data(filepath):
    """
    Reads an Excel or CSV file and returns its content as a list of lists.
    """
    if filepath.lower().endswith(('.xlsx', '.xls')):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        wb.close()
        return data
    elif filepath.lower().endswith('.csv'):
        with open(filepath, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            return list(reader)
    else:
        return None


def adjust_sheet_formatting(sheet):
    """
    Centers all text in each cell and adjusts the column widths with some padding.
    """
    # Iterate over all columns
    for col in sheet.columns:
        max_length = 0
        # Get the column letter from the first cell in the column.
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                cell_value = str(cell.value)
                # Center align the cell.
                cell.alignment = Alignment(horizontal="center")
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
        # Add a little padding.
        sheet.column_dimensions[column_letter].width = max_length + 4


def generate_average_files():
    folder1 = folder1_entry.get()
    folder2 = folder2_entry.get()
    save_folder = save_folder_entry.get()

    if not all([folder1, folder2, save_folder]):
        messagebox.showerror("Error", "Please select Folder 1, Folder 2, and the Save Folder.")
        return

    valid_ext = ('.xlsx', '.xls', '.csv')
    files1 = [f for f in os.listdir(folder1) if f.startswith("P") and f.lower().endswith(valid_ext)]
    files2 = [f for f in os.listdir(folder2) if f.startswith("P") and f.lower().endswith(valid_ext)]

    if not files1 or not files2:
        messagebox.showerror("Error", "No matching files found in one or both folders.")
        return

    # Compute a base key by removing the extension and then the last word.
    def get_base_key(filename):
        base = filename.rsplit('.', 1)[0]  # Remove extension.
        parts = base.split()
        if len(parts) > 1:
            return " ".join(parts[:-1])  # All parts except the last word.
        else:
            return base

    # Build dictionaries mapping base key to full file path.
    dict1 = {get_base_key(f): os.path.join(folder1, f) for f in files1}
    dict2 = {get_base_key(f): os.path.join(folder2, f) for f in files2}

    # Determine common base keys.
    common_keys = set(dict1.keys()).intersection(set(dict2.keys()))
    if not common_keys:
        messagebox.showerror("Error", "No matching file pairs found between the two folders.")
        return

    for base in sorted(common_keys):
        file1_path = dict1[base]
        file2_path = dict2[base]
        data1 = read_workbook_data(file1_path)
        data2 = read_workbook_data(file2_path)
        if data1 is None or data2 is None:
            messagebox.showerror("Error", f"Error reading files for base key '{base}'.")
            continue

        # Use the first file's structure as a template.
        num_rows = len(data1)
        num_cols = max(len(row) for row in data1)

        # Initialize accumulators for averaging (cells starting from B2: row index 1, col index 1).
        accumulator = [[0] * num_cols for _ in range(num_rows)]
        count_matrix = [[0] * num_cols for _ in range(num_rows)]

        # Loop over each data cell (skipping header row and first column)
        for r in range(1, num_rows):
            for c in range(1, num_cols):
                try:
                    val1 = float(data1[r][c]) if data1[r][c] not in (None, "") else 0
                except Exception:
                    val1 = 0
                try:
                    val2 = float(data2[r][c]) if data2[r][c] not in (None, "") else 0
                except Exception:
                    val2 = 0
                accumulator[r][c] = val1 + val2
                count_matrix[r][c] = 2  # Two values per pair

        # Create a new workbook for the averaged data.
        out_wb = openpyxl.Workbook()
        out_sheet = out_wb.active

        # Copy column A (headers/labels) from the template.
        for r in range(num_rows):
            val = data1[r][0] if len(data1[r]) > 0 else ""
            out_sheet.cell(row=r + 1, column=1, value=val)

        # Ensure cell B1 has "Value" if not already present.
        cell_b1 = out_sheet.cell(row=1, column=2)
        if not cell_b1.value or str(cell_b1.value).strip() == "":
            cell_b1.value = "Value"

        # Write averaged values into cells from B onward.
        for r in range(1, num_rows):
            for c in range(1, num_cols):
                avg_val = accumulator[r][c] / count_matrix[r][c] if count_matrix[r][c] > 0 else None
                out_sheet.cell(row=r + 1, column=c + 1, value=avg_val)

        # Adjust formatting: center text and auto-adjust column widths.
        adjust_sheet_formatting(out_sheet)

        # Create the output filename: e.g., "P1 Return Combined.xlsx"
        output_filename = f"{base} Combined.xlsx"
        output_path = os.path.join(save_folder, output_filename)
        try:
            out_wb.save(output_path)
        except Exception as e:
            messagebox.showerror("Error", f"Could not save file '{output_filename}': {e}")

    messagebox.showinfo("Success", "Averaged files have been generated and saved.")

    # Automatically open the save folder directory.
    try:
        if os.name == 'nt':
            os.startfile(save_folder)
        elif sys.platform == 'darwin':
            subprocess.call(["open", save_folder])
        else:
            subprocess.call(["xdg-open", save_folder])
    except Exception as e:
        messagebox.showerror("Error", f"Could not open directory: {e}")


# --------------------- Main GUI Setup --------------------- #
root = tk.Tk()
root.title("Pairwise Excel Averager")

style = ttk.Style(root)
style.theme_use("clam")
root.configure(bg='#f0f0f0')

pad_opts = {'padx': 5, 'pady': 5}

ttk.Label(root, text="Folder 1 (Condition 1):").grid(row=0, column=0, sticky="w", **pad_opts)
folder1_entry = ttk.Entry(root, width=50)
folder1_entry.grid(row=0, column=1, **pad_opts)
ttk.Button(root, text="Browse", command=select_folder1).grid(row=0, column=2, **pad_opts)

ttk.Label(root, text="Folder 2 (Condition 2):").grid(row=1, column=0, sticky="w", **pad_opts)
folder2_entry = ttk.Entry(root, width=50)
folder2_entry.grid(row=1, column=1, **pad_opts)
ttk.Button(root, text="Browse", command=select_folder2).grid(row=1, column=2, **pad_opts)

ttk.Label(root, text="Save Averaged Files To:").grid(row=2, column=0, sticky="w", **pad_opts)
save_folder_entry = ttk.Entry(root, width=50)
save_folder_entry.grid(row=2, column=1, **pad_opts)
ttk.Button(root, text="Browse", command=select_save_folder).grid(row=2, column=2, **pad_opts)

ttk.Button(root, text="Generate Averaged Files", command=generate_average_files) \
    .grid(row=3, column=1, pady=10)

root.mainloop()
